import sys
from awsglue.transforms import *
from awsglue.utils import getResolvedOptions
from pyspark.context import SparkContext
from awsglue.context import GlueContext
from awsglue.job import Job
from pyspark.sql.functions import col, lpad, explode
from awsglue import DynamicFrame
from awsglue.dynamicframe import DynamicFrameCollection
import datetime
import boto3
from io import BytesIO
import openpyxl
import pandas as pd
import logging


# Logging Setup
logger = logging.getLogger()
logger.setLevel(logging.INFO)
handler = logging.StreamHandler(sys.stdout)
handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
logger.addHandler(handler)

# Script generated for node Custom Transform
def MyTransform(glueContext, dfc) -> DynamicFrameCollection:
    dynamic_frame = dfc["AmazonKinesis_node"]
    df = dynamic_frame.toDF()

    if df.count() == 0:
        logger.info("No data in batch.")
        return DynamicFrameCollection({}, glueContext)

    try:
        df_unpacked = df.select(explode(col("numbers")).alias("number"))
        df_transformed = df_unpacked.withColumn(
            "number", lpad(col("number").cast("string"), 3, "0")
        )
        dynamic_frame_transformed = DynamicFrame.fromDF(df_transformed, glueContext, "transformed")
        logger.info(f"Transformed {df_transformed.count()} numbers in batch.")
        return DynamicFrameCollection({"transformed_number": dynamic_frame_transformed}, glueContext)
    except Exception as e:
        logger.error(f"Transform failed: {e}")
        return DynamicFrameCollection({}, glueContext)


args = getResolvedOptions(sys.argv, ['JOB_NAME'])
sc = SparkContext()
glueContext = GlueContext(sc)
spark = glueContext.spark_session
job = Job(glueContext)
job.init(args['JOB_NAME'], args)

# Script generated for node Amazon Kinesis
dataframe_AmazonKinesis_node1741124607807 = glueContext.create_data_frame.from_options(connection_type="kinesis",connection_options={"typeOfData": "kinesis", "streamARN": "arn:aws:kinesis:us-east-1:851725381788:stream/coffeeStream", "classification": "json", "startingPosition": "earliest", "inferSchema": "true"}, transformation_ctx="dataframe_AmazonKinesis_node1741124607807")

# Excel Save Function
def save_to_excel_and_upload_to_s3(df, s3_base_path):
    try:
        workbook_count = 1
        sheet_count = 1
        row = 1
        col = 1
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"Z{sheet_count}"

        for idx, row_data in df.iterrows():
            for value in row_data:
                sheet.cell(row=row, column=col, value=value)
                col += 1
                if col > 1000:
                    col = 1
                    row += 1
            if row > 1000:
                if sheet_count >= 1000:
                    file_stream = BytesIO()
                    workbook.save(file_stream)
                    file_stream.seek(0)
                    s3_path = f"{s3_base_path}_workbook_{workbook_count}.xlsx"
                    s3_client = boto3.client('s3')
                    s3_client.put_object(Body=file_stream, Bucket="my-bucket-sun-test", Key=s3_path)
                    logger.info(f"Uploaded {s3_path}")
                    workbook_count += 1
                    workbook = openpyxl.Workbook()  # Create new workbook
                    sheet = workbook.active        # Use the active sheet
                    sheet_count = 1                # Reset sheet count
                    sheet.title = f"Z{sheet_count}"  # Set title to Z1
                    row = 1
                    col = 1
                else:
                    sheet_count += 1
                    row = 1
                    col = 1
                    sheet = workbook.create_sheet(title=f"Z{sheet_count}")

        if row > 1 or col > 1:
            file_stream = BytesIO()
            workbook.save(file_stream)
            file_stream.seek(0)
            s3_path = f"{s3_base_path}_workbook_{workbook_count}.xlsx"
            s3_client = boto3.client('s3')
            s3_client.put_object(Body=file_stream, Bucket="my-bucket-sun-test", Key=s3_path)
            logger.info(f"Uploaded {s3_path}")
    except Exception as e:
        logger.error(f"Excel save/upload failed: {e}")
# Batch Processing
def processBatch(data_frame, batchId):
    try:
        if data_frame.count() > 0:
            dynamic_frame = DynamicFrame.fromDF(data_frame, glueContext, "from_data_frame")
            transform_result = MyTransform(glueContext, DynamicFrameCollection({"AmazonKinesis_node": dynamic_frame}, glueContext))
            transformed_dynamic_frame = transform_result["transformed_number"]
            transformed_df = transformed_dynamic_frame.toDF()
            transformed_df_pandas = transformed_df.toPandas()

            now = datetime.datetime.now()
            s3_base_path = f"s3://my-bucket-sun-test/temp/ingest_year={now.year:0>4}/ingest_month={now.month:0>2}/ingest_day={now.day:0>2}/ingest_hour={now.hour:0>2}/output"
            save_to_excel_and_upload_to_s3(transformed_df_pandas, s3_base_path)
            logger.info(f"Processed batch {batchId} with {transformed_df.count()} numbers")
        else:
            logger.info(f"Batch {batchId} is empty")
    except Exception as e:
        logger.error(f"Batch {batchId} processing failed: {e}")

# Run for 1 Hour
logger.info("Starting Glue job for 1 hour...")

glueContext.forEachBatch(frame = dataframe_AmazonKinesis_node1741124607807, batch_function = processBatch, options = {"windowSize": "60 seconds", "checkpointLocation": args["TempDir"] + "/" + args["JOB_NAME"] + "/checkpoint/"})
job.commit()
logger.info("Glue job completed.")