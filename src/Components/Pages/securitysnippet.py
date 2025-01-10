import openpyxl

# Open the workbook with 'read_only' mode to save memory
wb = openpyxl.load_workbook('your_workbook.xlsx', read_only=True)

# Open a text file to write the results
with open('output.txt', 'w') as f:
    # Iterate over the first 10 sheets (or as many as exist)
    for sheet_index, sheet_name in enumerate(wb.sheetnames[:10]):
        sheet = wb[sheet_name]
        
        # Try to get data from the cell at row 288, column 600
        try:
            # row 288 and column 600 correspond to (288, 600) in openpyxl (1-based index)
            cell_value = sheet.cell(row=288, column=600).value
            
            # If no data is found, append '000'
            if cell_value is None:
                cell_value = '000'
            
            # Write the sheet name/number and the value to the output file
            f.write(f"Sheet {sheet_index + 1} ({sheet_name}): {cell_value}\n")
        
        except Exception as e:
            # In case the row/column is out of range or another error occurs, write an error message
            f.write(f"Error in Sheet {sheet_index + 1} ({sheet_name}): {str(e)}\n")
