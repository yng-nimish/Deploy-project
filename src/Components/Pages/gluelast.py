import sys
from awsglue.transforms import *
from awsglue.utils import getResolvedOptions
from pyspark.context import SparkContext
from awsglue.context import GlueContext
from awsglue.job import Job
from pyspark.sql.functions import col, lpad, explode, count
from awsglue import DynamicFrame
from awsglue.dynamicframe import DynamicFrameCollection
import datetime
import boto3
from io import BytesIO
import openpyxl
import logging

# Logging Setup
logger = logging.getLogger()
logger.setLevel(logging.INFO)
handler = logging.StreamHandler(sys.stdout)
handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
logger.addHandler(handler)

# Get job arguments
args = getResolvedOptions(sys.argv, ['JOB_NAME', 'TempDir'])

# Initialize GlueContext
glueContext = GlueContext(SparkContext.getOrCreate())
spark = glueContext.spark_session

spark.conf.set("spark.sql.execution.arrow.enabled", "true")
logger.info("Arrow optimization enabled for PySpark to Pandas conversion.")

job = Job(glueContext)
job.init(args['JOB_NAME'], args)

def MyTransform(glueContext, dfc) -> DynamicFrameCollection:
    dynamic_frame = dfc["AmazonKinesis_node"]
    df = dynamic_frame.toDF()  # No cache
    record_count = df.count()
    logger.info(f"MyTransform: Received {record_count} records")

    if record_count == 0:
        logger.info("MyTransform: No data in batch.")
        return DynamicFrameCollection({}, glueContext)

    try:
        df.printSchema()
        logger.info(f"MyTransform: Input schema logged")
        df_unpacked = df.select(explode(col("numbers")).alias("number"))
        transformed_count = df_unpacked.count()
        df_transformed = df_unpacked.withColumn(
            "number", lpad(col("number").cast("string"), 3, "0")
        )
        dynamic_frame_transformed = DynamicFrame.fromDF(df_transformed, glueContext, "transformed")
        logger.info(f"MyTransform: Transformed {transformed_count} numbers from {record_count} records")
        return DynamicFrameCollection({"transformed_number": dynamic_frame_transformed}, glueContext)
    except Exception as e:
        logger.error(f"MyTransform: Transform failed: {e}")
        return DynamicFrameCollection({}, glueContext)

def save_to_excel_and_upload_to_s3(df, s3_base_path):
    logger.info("Entered excel Function 1")
    try:
        total_rows = df.count()
        logger.info(f"Total numbers to process: {total_rows}")

        workbook_count = 1
        workbook = openpyxl.Workbook()
        sheet_count = 0
        row = 1
        col = 1

        workbook.remove(workbook.active)
        sheet = workbook.create_sheet(title=f"Z{sheet_count + 1}")
        sheet_count += 1

        numbers = df.select("number").rdd.flatMap(lambda x: x).collect()
        logger.info(f"Collected {len(numbers)} numbers from DataFrame")

        for idx, value in enumerate(numbers):
            sheet.cell(row=row, column=col, value=value)
            col += 1

            if col > 1000:
                col = 1
                row += 1

            if row > 1000:
                # Save and upload after each sheet is filled (including the first)
                file_stream = BytesIO()
                workbook.save(file_stream)
                file_stream.seek(0)
                s3_path = f"{s3_base_path}_workbook_{workbook_count}.xlsx"
                s3_client = boto3.client('s3')
                s3_client.put_object(Body=file_stream, Bucket="my-bucket-sun-test", Key=s3_path)
                logger.info(f"Uploaded {s3_path}")
                workbook_count += 1

                # Start a new workbook
                workbook = openpyxl.Workbook()
                workbook.remove(workbook.active)
                if idx + 1 < total_rows:  # Only create new sheet if more data remains
                    sheet = workbook.create_sheet(title=f"Z{sheet_count + 1}")
                    sheet_count += 1
                    row = 1
                    col = 1

        # Save any remaining partial sheet
        if row > 1 or col > 1:
            file_stream = BytesIO()
            workbook.save(file_stream)
            file_stream.seek(0)
            s3_path = f"{s3_base_path}_workbook_{workbook_count}.xlsx"
            s3_client = boto3.client('s3')
            s3_client.put_object(Body=file_stream, Bucket="my-bucket-sun-test", Key=s3_path)
            logger.info(f"Uploaded {s3_path}")

        logger.info(f"Completed saving {total_rows} numbers across {workbook_count} workbooks")

    except Exception as e:
        logger.error(f"Excel save/upload failed: {e}")
        raise

def processBatch(data_frame, batchId):
    try:
        logger.info(f"Batch {batchId}: Starting processing")
        initial_count = data_frame.count()
        if initial_count > 0:
            logger.info(f"Batch {batchId}: Fetched {initial_count} records from Kinesis")
            dynamic_frame = DynamicFrame.fromDF(data_frame, glueContext, "from_data_frame")
            transform_result = MyTransform(glueContext, DynamicFrameCollection({"AmazonKinesis_node": dynamic_frame}, glueContext))
            
            if "transformed_number" not in transform_result.keys():
                logger.info(f"Batch {batchId}: No transformed data to process.")
                return

            transformed_dynamic_frame = transform_result["transformed_number"]
            transformed_df = transformed_dynamic_frame.toDF()
            transformed_df.show(7)
            
            total_numbers = transformed_df.count()
            logger.info(f"Batch {batchId}: Contains {total_numbers} transformed numbers")

            transformed_df.show(8)  # No repartitioning

            now = datetime.datetime.now()
            s3_base_path = f"s3://my-bucket-sun-test/temp/ingest_year={now.year:0>4}/ingest_month={now.month:0>2}/ingest_day={now.day:0>2}/output_batch_{batchId}"

            logger.info(f"Batch {batchId}: Checking number distribution before saving")
            transformed_df.groupBy("number").agg(count("number").alias("count")).orderBy(col("count").desc()).show(10)

            save_to_excel_and_upload_to_s3(transformed_df, s3_base_path)
            logger.info(f"Processed batch {batchId} with {transformed_df.count()} numbers")
        else:
            logger.info(f"Batch {batchId}: No data fetched from Kinesis")
    except Exception as e:
        logger.error(f"Batch {batchId} processing failed: {e}")
        raise

# Script generated for node Amazon Kinesis
dataframe_AmazonKinesis_node1743426800728 = glueContext.create_data_frame.from_options(
    connection_type="kinesis",
    connection_options={
        "typeOfData": "kinesis",
        "streamARN": "arn:aws:kinesis:us-east-1:851725381788:stream/ExpertStream",
        "classification": "json",
        "startingPosition": "earliest",
        "inferSchema": "true"
    },
    transformation_ctx="dataframe_AmazonKinesis_node1743426800728"
)

glueContext.forEachBatch(
    frame=dataframe_AmazonKinesis_node1743426800728,
    batch_function=processBatch,
    options={"windowSize": "100 seconds", "checkpointLocation": args["TempDir"] + "/" + args["JOB_NAME"] + "/checkpoint/"}
)
job.commit()
logger.info("Glue job completed.")