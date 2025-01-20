import sys
import os
import pandas as pd
from awsglue.transforms import *
from awsglue.utils import getResolvedOptions
from pyspark.context import SparkContext
from awsglue.context import GlueContext
from awsglue.job import Job
from pyspark.sql.functions import col, lpad
from pyspark.sql import DataFrame, Row
import datetime
from awsglue import DynamicFrame
from awsglue.dynamicframe import DynamicFrameCollection
import openpyxl  # Ensure openpyxl is imported
import boto3
from io import BytesIO

# Script generated for node Custom Transform
def MyTransform(glueContext, dfc) -> DynamicFrameCollection:
    dynamic_frame = dfc["AmazonKinesis_node1736967491930"]
    
    # Convert to DataFrame
    df = dynamic_frame.toDF()

    # Check if DataFrame is empty
    if df.count() == 0:
        print("No data to process in this batch.")
        return DynamicFrameCollection({}, glueContext)
    
    # Check if the 'number' column exists
    if "number" not in df.columns:
        print("Column 'number' not found. Skipping transformation.")
        return DynamicFrameCollection({}, glueContext)

    # Perform the transformation: Pad 'number' column to 3 digits
    df_transformed = df.withColumn(
        "number",
        lpad(col("number").cast("string"), 3, "0")
    )
    
    # Convert back to DynamicFrame
    dynamic_frame_transformed = DynamicFrame.fromDF(df_transformed, glueContext, "dynamic_frame_transformed")

    # Return the transformed DynamicFrameCollection
    return DynamicFrameCollection({"transformed_number": dynamic_frame_transformed}, glueContext)


# Getting the job arguments
args = getResolvedOptions(sys.argv, ['JOB_NAME', 'TempDir'])

# Initialize Glue context
sc = SparkContext()
glueContext = GlueContext(sc)
spark = glueContext.spark_session
job = Job(glueContext)
job.init(args['JOB_NAME'], args)

# Script generated for node Amazon Kinesis
dataframe_AmazonKinesis_node1736967491930 = glueContext.create_data_frame.from_options(
    connection_type="kinesis",
    connection_options={"typeOfData": "kinesis", "streamARN": "arn:aws:kinesis:us-east-1:851725381788:stream/River", "classification": "json", "startingPosition": "earliest", "inferSchema": "true"},
    transformation_ctx="dataframe_AmazonKinesis_node1736967491930"
)

# Function to create Excel file and save to S3
def save_to_excel_and_upload_to_s3(df, s3_path):
    # Create a new Excel workbook and set up sheet parameters
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    
    # Iterate through the DataFrame and write data to Excel
    row = 1
    col = 1
    for idx, row_data in df.iterrows():
        for value in row_data:
            sheet.cell(row=row, column=col, value=value)
            col += 1
            if col > 100:
                col = 1
                row += 1
        if row > 100:
            sheet = workbook.create_sheet(title=f"Sheet{workbook.sheetnames[-1][-1]}")
            row = 1
            col = 1
    
    # Save the workbook to a BytesIO object to avoid file I/O
    file_stream = BytesIO()
    workbook.save(file_stream)
    file_stream.seek(0)
    
    # Upload the file to S3
    s3_client = boto3.client('s3')
    s3_client.put_object(
        Body=file_stream,
        Bucket="my-bucket-sun-test",
        Key=s3_path
    )

def processBatch(data_frame, batchId):
    if (data_frame.count() > 0):
        AmazonKinesis_node1736967491930 = DynamicFrame.fromDF(data_frame, glueContext, "from_data_frame")
        
        # Script generated for node Custom Transform
        CustomTransform_node1736967498329 = MyTransform(glueContext, {"AmazonKinesis_node1736967491930": AmazonKinesis_node1736967491930})

        # Access the transformed DynamicFrame from the collection
        transformed_dynamic_frame = CustomTransform_node1736967498329["transformed_number"]
        
        # Convert the transformed DynamicFrame to a DataFrame
        transformed_df = transformed_dynamic_frame.toDF()

        # Convert to Pandas DataFrame for Excel export
        transformed_df_pandas = transformed_df.toPandas()

        # Generate current date and time for partitioning
        now = datetime.datetime.now()
        year = now.year
        month = now.month
        day = now.day
        hour = now.hour
        # S3 Path for saving Excel
        s3_path = f"s3://my-bucket-sun-test/temp/ingest_year={year:0>4}/ingest_month={month:0>2}/ingest_day={day:0>2}/ingest_hour={hour:0>2}/output.xlsx"
        
        # Save the transformed DataFrame to Excel and upload to S3
        save_to_excel_and_upload_to_s3(transformed_df_pandas, s3_path)

# Process the data in batches
glueContext.forEachBatch(
    frame=dataframe_AmazonKinesis_node1736967491930,
    batch_function=processBatch,
    options={"windowSize": "10 seconds", "checkpointLocation": args["TempDir"] + "/" + args["JOB_NAME"] + "/checkpoint/"}
)

# Commit the Glue job
job.commit()
